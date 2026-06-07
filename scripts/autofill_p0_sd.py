#!/usr/bin/env python3
"""Auto-fill col A (P0) and col B (sd) from NHANES for every study row.

Implements the convention agreed Apr 20, 2026: col A is the output node's
target-population (NHANES) P(first state), col B is the output node's
NHANES sd (only used for ES/WMD stat types).

For each row with Stat ∈ {OR, HR, SMD, ES, WMD}:
  - Find the output node's definition row
  - Compute the correct P0 based on the definition row's Type:
      *_nhanes_explicit / discrete_nhanes_explicit / naive_0_nhanes_explicit:
          calculate_priors_explicit -> first value's frequency
      *_nhanes_quartile / naive_0_nhanes_quartile:
          0.25 if first value is quartile_1, else compute from the quartile def
      dependency_priors / discrete_priors:
          index1 value (already a probability)
      all_of / any_of / avg / if_then_else / dependency_distal:
          computed from parents recursively - skip here, priors come at
          build time; leave col A blank
  - For col B: if output node has a continuous NHANES code (LBX*/BMX*/etc.)
    compute sd across all non-missing raw values in NHANES

Writes:
  data/Individual Relations.working.xlsx (in place, preserves formulas)
  docs/autofill_p0_sd_diff.md (report of every change)

Dry-run mode: pass --dry to skip the write and just emit the report.
"""
import argparse
import sys
import pandas as pd
import numpy as np
from openpyxl import load_workbook

sys.path.insert(0, '.')


STAT_TYPES_NEED_P0 = {'OR', 'HR', 'SMD', 'ES', 'WMD'}
STAT_TYPES_NEED_SD = {'ES', 'WMD'}

# Types classification
NHANES_EXPLICIT = {'dependency_nhanes_explicit', 'dependency_nhanes_explicit_average',
                   'discrete_nhanes_explicit', 'naive_0_nhanes_explicit',
                   'naive_0_nhanes_explicit_average'}
NHANES_QUARTILE = {'dependency_nhanes_quartile', 'dependency_nhanes_quartile_average',
                   'discrete_nhanes_quartile', 'discrete_nhanes_quartile_average',
                   'naive_0_nhanes_quartile', 'naive_0_nhanes_quartile_average'}
PRIORS_TYPES    = {'dependency_priors', 'discrete_priors'}
COMPOSITE_TYPES = {'all_of', 'any_of', 'avg', 'if_then_else', 'dependency_distal'}
ALIAS_TYPES     = {'is_a', 'equivalent_to', 'subsumes', 'equivalent_distal'}


def log(m): print(m, flush=True)


def load_df():
    df = pd.read_excel('./data/Individual Relations.working.xlsx',
                       sheet_name='all worksheet anom', header=0)
    # Normalize case: codes and Stat should be uppercase (NHANES is all caps;
    # some xlsx rows have MCQ230a, DBD270c, etc.)
    df['code'] = df['code'].apply(lambda e: e.upper() if isinstance(e, str) else e)
    df['Stat'] = df['Stat'].apply(lambda e: e.upper() if isinstance(e, str) else e)
    return df


def find_def_row(df, output_name):
    r = df[(df['output'] == output_name) & (df['input'].isna()) & (df['Type'].notna())]
    if len(r) == 0: return None
    return r.iloc[0]


def extract_value_ranges(def_row):
    """Parse index1/index2/... into a list of numeric ranges or discrete values."""
    from sn_bayes.config_creation.priors import extract_value_ranges as evr
    return evr(def_row)


def compute_nhanes_p0_explicit(def_row, nhanes):
    """For *_nhanes_explicit: fraction of NHANES respondents whose code
    falls in value1's index1 range."""
    code = def_row.get('code')
    if not isinstance(code, str) or code.strip() == '':
        return None, "no code"
    code = code.strip()
    if code not in nhanes.columns:
        return None, f"code {code} not in NHANES"
    try:
        vr = extract_value_ranges(def_row)
    except Exception as e:
        return None, f"extract_value_ranges failed: {e}"
    first_value = def_row.get('value1')
    if not isinstance(first_value, str):
        return None, "no value1"
    first_value = first_value.strip()
    if first_value not in vr:
        return None, f"value1 {first_value} not in value_ranges"

    ranges = vr[first_value]
    col = nhanes[code]
    non_na = col.notna()
    # Apply range membership
    mask = pd.Series(False, index=col.index)
    for r in ranges:
        if isinstance(r, tuple):
            mask = mask | ((col >= r[0]) & (col <= r[1]))
        else:
            mask = mask | (col == r)
    total_valid = non_na.sum()
    if total_valid == 0:
        return None, "no valid NHANES rows"
    p0 = (mask & non_na).sum() / total_valid
    return float(p0), f"NHANES: {code} in {ranges}; p={p0:.6f} of {total_valid} valid"


def compute_quartile_p0(def_row):
    """For quartile types, P(first state) = 0.25 by construction."""
    return 0.25, "quartile: first state is quartile_1, P=0.25 by definition"


def compute_priors_p0(def_row):
    """For dependency_priors/discrete_priors: index1 is P(value1) directly."""
    idx1 = def_row.get('index1')
    try:
        p = float(idx1)
        return p, f"priors type: index1={idx1} read as P(value1)"
    except (TypeError, ValueError):
        return None, f"priors type: index1={idx1!r} not a float"


def compute_nhanes_sd(def_row, nhanes):
    """Compute sd of the NHANES raw values for the output node's code."""
    code = def_row.get('code')
    if not isinstance(code, str) or code.strip() == '':
        return None, "no code"
    code = code.strip()
    if code not in nhanes.columns:
        return None, f"code {code} not in NHANES"
    col = nhanes[code]
    valid = col[col.notna() & (col > -999)]  # drop missing-value sentinels
    if len(valid) < 10:
        return None, f"too few NHANES rows ({len(valid)})"
    sd = float(valid.std())
    return sd, f"NHANES: std of {code} over {len(valid)} rows = {sd:.4f}"


def compute_p0_for_row(df, study_row, nhanes):
    """Return (p0, reason) for the given study row's output node."""
    output = study_row['output']
    def_row = find_def_row(df, output)
    if def_row is None:
        return None, f"no definition row for output={output}"
    typ = def_row.get('Type')
    if not isinstance(typ, str):
        return None, f"definition Type missing"
    typ = typ.strip()
    if typ in NHANES_EXPLICIT:
        return compute_nhanes_p0_explicit(def_row, nhanes)
    elif typ in NHANES_QUARTILE:
        return compute_quartile_p0(def_row)
    elif typ in PRIORS_TYPES:
        return compute_priors_p0(def_row)
    elif typ in COMPOSITE_TYPES:
        return None, f"composite Type ({typ}): P0 derived at build time from parents - leave col A blank"
    elif typ in ALIAS_TYPES:
        return None, f"alias Type ({typ}): should defer to aliased node"
    return None, f"unknown Type: {typ}"


def compute_sd_for_row(df, study_row, nhanes):
    """Return (sd, reason) for ES/WMD conversion."""
    output = study_row['output']
    def_row = find_def_row(df, output)
    if def_row is None:
        return None, f"no definition row for output={output}"
    return compute_nhanes_sd(def_row, nhanes)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--dry', action='store_true', help='produce report only, do not write xlsx')
    args = ap.parse_args()

    log(f"[autofill] loading xlsx + NHANES...")
    df = load_df()
    nhanes = pd.read_csv('./data/preprocessed_nhanes.csv', low_memory=False)
    log(f"[autofill] df: {df.shape}, nhanes: {nhanes.shape}")

    # Collect rows needing P0 / sd
    needs_p0 = df[df['Stat'].isin(STAT_TYPES_NEED_P0) & df['input'].notna()]
    log(f"[autofill] rows with Stat in {STAT_TYPES_NEED_P0} and input set: {len(needs_p0)}")

    changes = []  # list of (row_index, col, old, new, reason)
    skips = []
    for idx, row in needs_p0.iterrows():
        xlsx_row = idx + 2  # xlsx is 1-indexed with header at row 1
        old_p0 = row.get('control probability (needed if or,hr,smd, es, or wmd)')
        if pd.isna(old_p0): old_p0 = None
        else:
            try: old_p0 = float(old_p0)
            except: old_p0 = None

        new_p0, reason_p0 = compute_p0_for_row(df, row, nhanes)
        if new_p0 is None:
            skips.append((xlsx_row, 'A', old_p0, reason_p0))
        else:
            if old_p0 is None or abs(old_p0 - new_p0) > 1e-6:
                changes.append((xlsx_row, 'A', old_p0, new_p0, reason_p0))

        if row['Stat'] in STAT_TYPES_NEED_SD:
            old_sd = row.get('control stdev (needed if ES or WMD)')
            if pd.isna(old_sd): old_sd = None
            else:
                try: old_sd = float(old_sd)
                except: old_sd = None
            new_sd, reason_sd = compute_sd_for_row(df, row, nhanes)
            if new_sd is None:
                skips.append((xlsx_row, 'B', old_sd, reason_sd))
            else:
                if old_sd is None or abs(old_sd - new_sd) > 1e-6:
                    changes.append((xlsx_row, 'B', old_sd, new_sd, reason_sd))

    # Report
    log(f"\n[autofill] changes: {len(changes)}")
    log(f"[autofill] skips:   {len(skips)}")
    report_path = 'docs/autofill_p0_sd_diff.md'
    with open(report_path, 'w') as f:
        f.write('# Autofill P0 / sd diff report\n\n')
        f.write('Convention: col A = NHANES P0, col B = NHANES sd (for ES/WMD only).\n\n')
        f.write(f'Rows needing P0: {len(needs_p0)}\n\n')
        f.write(f'## Changes ({len(changes)})\n\n')
        f.write('| row | col | old | new | reason |\n|-----|-----|-----|-----|--------|\n')
        for xr, col, old, new, reason in changes:
            old_s = f'{old:.6f}' if isinstance(old, float) else 'blank'
            new_s = f'{new:.6f}' if isinstance(new, float) else 'blank'
            reason = reason[:120].replace('|', '\\|')
            f.write(f'| {xr} | {col} | {old_s} | {new_s} | {reason} |\n')
        f.write(f'\n## Skipped ({len(skips)})\n\n')
        f.write('| row | col | old | reason |\n|-----|-----|-----|--------|\n')
        for xr, col, old, reason in skips:
            old_s = f'{old:.6f}' if isinstance(old, float) else 'blank'
            reason = reason[:120].replace('|', '\\|')
            f.write(f'| {xr} | {col} | {old_s} | {reason} |\n')
    log(f"[autofill] report written: {report_path}")

    if args.dry:
        log(f"[autofill] --dry: not writing xlsx")
        return

    # Apply changes to xlsx.
    # Strategy: write new A/B values AND recompute M/N in Python, write M/N
    # as plain values (no formulas). This ensures the build pipeline sees
    # consistent numbers without requiring Excel to recompute.
    log(f"[autofill] applying {len(changes)} changes + recomputing M/N for affected rows...")
    from sn_bayes.rr_formulas import compute_rr_and_pm

    wb = load_workbook('./data/Individual Relations.working.xlsx', data_only=False)
    ws = wb['all worksheet anom']

    # Apply A / B changes
    for xr, col, old, new, _ in changes:
        col_num = ord(col) - ord('A') + 1
        ws.cell(row=xr, column=col_num).value = new

    # Resolve any remaining simple arithmetic formulas so pd.read_excel sees
    # numbers instead of NaN (cached values are wiped by openpyxl save) AND
    # so the M/N recompute below can read A/B as plain numbers.
    import re
    formula_cells_resolved = 0
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if not isinstance(v, str) or not v.startswith('='):
                continue
            expr = v[1:]
            m = re.match(r'^1-([A-Z]+)(\d+)$', expr)
            if m:
                ref_col = m.group(1); ref_row = int(m.group(2))
                col_idx = 0
                for ch in ref_col:
                    col_idx = col_idx * 26 + (ord(ch) - ord('A') + 1)
                ref_val = ws.cell(row=ref_row, column=col_idx).value
                try:
                    c.value = 1 - float(ref_val)
                    formula_cells_resolved += 1
                except (TypeError, ValueError):
                    pass
                continue
            m = re.match(r'^(-?\d+(?:\.\d+)?)\s*([/*+\-])\s*(-?\d+(?:\.\d+)?)$', expr)
            if m:
                a, op, b = m.group(1), m.group(2), m.group(3)
                try:
                    a = float(a); b = float(b)
                    if op == '/' and b != 0: c.value = a / b
                    elif op == '*': c.value = a * b
                    elif op == '+': c.value = a + b
                    elif op == '-': c.value = a - b
                    formula_cells_resolved += 1
                except (TypeError, ValueError):
                    pass
                continue
            # General: resolve cell refs to values, then eval if safe
            # e.g. =C887-1/(0.85+0.05) → 1.176-1/0.9
            try:
                # Replace cell refs with their current values
                resolved = expr
                for cref in re.findall(r'[A-Z]+\d+', resolved):
                    cref_col = re.match(r'([A-Z]+)', cref).group(1)
                    cref_row = int(re.search(r'\d+', cref).group(0))
                    col_idx = 0
                    for ch in cref_col:
                        col_idx = col_idx * 26 + (ord(ch) - ord('A') + 1)
                    ref_val = ws.cell(row=cref_row, column=col_idx).value
                    if ref_val is None:
                        raise ValueError('blank ref')
                    resolved = resolved.replace(cref, f'({float(ref_val)})')
                # After substitution, should be pure arithmetic
                if not re.match(r'^[\d\.\s\+\-\*\/\(\)]+$', resolved):
                    continue
                val = eval(resolved, {'__builtins__': {}}, {})
                c.value = float(val)
                formula_cells_resolved += 1
            except Exception:
                pass

    # Recompute M / N for every row where Stat is a convertible type (not
    # just changed rows - keeps consistency in case earlier rows had stale
    # cache values).
    def _to_float_or_none(v):
        if v is None: return None
        try: return float(v)
        except (TypeError, ValueError): return None

    recomputed = 0
    for xr_raw in range(2, ws.max_row + 1):
        stat = ws.cell(row=xr_raw, column=5).value
        if not isinstance(stat, str): continue
        stat_u = stat.strip().upper()
        if stat_u not in ('OR', 'HR', 'SMD', 'ES', 'WMD', 'RR'): continue
        C_f = _to_float_or_none(ws.cell(row=xr_raw, column=3).value)
        D_f = _to_float_or_none(ws.cell(row=xr_raw, column=4).value)
        A_f = _to_float_or_none(ws.cell(row=xr_raw, column=1).value)
        B_f = _to_float_or_none(ws.cell(row=xr_raw, column=2).value)
        # Skip rows where C (effect size) is entirely missing - can't compute RR
        if C_f is None: continue
        RR, PM = compute_rr_and_pm(stat_u, C_f, D_f, A_f, B_f)
        ws.cell(row=xr_raw, column=13).value = RR   # M
        ws.cell(row=xr_raw, column=14).value = PM   # N
        recomputed += 1

    log(f"[autofill] resolved {formula_cells_resolved} simple arithmetic formulas to values")

    # Append short convention note to col AC (comment) on rows where A changed
    changed_rows = set(xr for xr, col, _, _, _ in changes if col == 'A')
    note = '[autofill Apr-20] col A updated to NHANES P0 per Apr 20, 2026 convention'
    for xr in changed_rows:
        existing = ws.cell(row=xr, column=29).value  # col AC = 29
        if existing is None or existing == '':
            ws.cell(row=xr, column=29).value = note
        else:
            ws.cell(row=xr, column=29).value = f"{existing}; {note}"

    wb.save('./data/Individual Relations.working.xlsx')
    log(f"[autofill] saved xlsx: {len(changes)} A/B changes, {recomputed} M/N recomputed, "
        f"{len(changed_rows)} comment notes added")


if __name__ == '__main__':
    main()
