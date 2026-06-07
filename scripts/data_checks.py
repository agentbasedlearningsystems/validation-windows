#!/usr/bin/env python3
"""
Data checks at every stage of the BayesExpert pipeline.
Run the appropriate check at each stage. All must pass before proceeding.

Usage:
    python3 data_checks.py spreadsheet              # After spreadsheet edit, before config
    python3 data_checks.py config <config.json>      # After config creation, before build
    python3 data_checks.py pickle <pickle_path>      # After build, before query
    python3 data_checks.py post_run <pickle> <config> # After full run

Options:
    --xlsx <path>       Override spreadsheet path (default: ./data/Individual Relations.working.xlsx)
    --sheet <name>      Override sheet name (default: all worksheet anom)
    --build-dir <path>  Override build directory
"""
import sys
import os
import json
import pickle
import math
import warnings
import numpy as np
import pandas as pd
from sn_bayes.utils import smart_load_pickle
warnings.filterwarnings('ignore')

XLSX = './data/Individual Relations.working.xlsx'
CSV_PATH = './data/relations.csv'
SHEET = 'all worksheet anom'
FORMULA_BASELINE = {'M': 1080, 'N': 1206, 'S': 0}  # Apr 20 2026: autofill_p0_sd.py replaced M/N/S cell formulas with plain values on rows where Stat Value was set; remaining M-formulas (1080) and N-formulas (1206) are rows with blank Stat where the formula evaluates to "" - harmless placeholders. S-formulas (dependency_priors index2 = 1-index1) are fully resolved to values.
BUILDS_DIR = './builds'
USE_CSV = True  # CSV is canonical post-migration (tasks #106-#109). Use --xlsx to override.

def get_latest_build_dir():
    """Find most recent build directory."""
    if not os.path.isdir(BUILDS_DIR): return None
    dirs = [d for d in sorted(os.listdir(BUILDS_DIR)) if os.path.isdir(os.path.join(BUILDS_DIR, d)) and d[0] == '2']
    return os.path.join(BUILDS_DIR, dirs[-1]) if dirs else None

def load_previous_counts():
    """Load DD/CPT counts from most recent build's data_checks.json."""
    build_dir = get_latest_build_dir()
    if not build_dir: return None
    check_file = os.path.join(build_dir, 'data_checks.json')
    if not os.path.exists(check_file): return None
    with open(check_file) as f:
        return json.load(f)

def save_counts(build_dir, counts):
    """Save DD/CPT counts to build directory."""
    if build_dir:
        os.makedirs(build_dir, exist_ok=True)
        with open(os.path.join(build_dir, 'data_checks.json'), 'w') as f:
            json.dump(counts, f, indent=2)


def check_spreadsheet():
    """Run before config creation. Checks formulas, data quality, structure."""
    import openpyxl
    import re

    print('=' * 60)
    print(f'CHECK 1: {"CSV" if USE_CSV else "SPREADSHEET"} (before config creation)')
    print('=' * 60)
    issues = 0

    if USE_CSV:
        data = pd.read_csv(CSV_PATH)
        is_default_xlsx = False
        print(f'  Source: {CSV_PATH}')
        print(f'  Formula checks: SKIPPED (CSV has no formulas - by design after migration)')
    else:
        wb = openpyxl.load_workbook(XLSX)
        ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active
        fc = {'M': 0, 'N': 0, 'S': 0}
        cross_row_cells = []
        col_templates = {'M': {}, 'N': {}, 'S': {}}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            rn = row[0].row
            for col, ci in [('M', 12), ('N', 13), ('S', 18)]:
                c = row[ci]
                if c.value and isinstance(c.value, str) and c.value.startswith('='):
                    fc[col] += 1
                    for r in re.findall(r'[A-Z]+(\d+)', c.value):
                        if r != str(rn):
                            cross_row_cells.append((col, rn, r))
                            break
                    template = re.sub(rf'(?<![0-9])({rn})(?![0-9])', '%ROW%', c.value)
                    template = re.sub(r'\s+', ' ', template).strip()
                    col_templates[col].setdefault(template, []).append(rn)

        is_default_xlsx = os.path.basename(XLSX) == 'Individual Relations.working.xlsx'
        if is_default_xlsx:
            for col in ['M', 'N', 'S']:
                ok = fc[col] == FORMULA_BASELINE[col]
                status = 'PASS' if ok else 'FAIL'
                print(f'  {col} formulas: {fc[col]} (baseline {FORMULA_BASELINE[col]}) {status}')
                if not ok: issues += 1
        else:
            print(f'  Formula check: SKIPPED (non-default xlsx)')

        CROSS_ROW_BASELINE = 6
        n_cross = len(cross_row_cells)
        if is_default_xlsx:
            cr_status = 'PASS' if n_cross <= CROSS_ROW_BASELINE else f'FAIL (>{CROSS_ROW_BASELINE} allowed)'
            print(f'  Own-row rule (cells w/ cross-row refs): {n_cross} (baseline ≤{CROSS_ROW_BASELINE}) {cr_status}')
            if n_cross > CROSS_ROW_BASELINE:
                for col, rn, other in cross_row_cells[:10]:
                    print(f'    - col {col} row {rn} refs row {other}')
                if n_cross > 10:
                    print(f'    ... +{n_cross-10} more')
                issues += 1
        else:
            print(f'  Own-row rule (cells w/ cross-row refs): {n_cross}')

        data = pd.read_excel(XLSX, sheet_name=SHEET if SHEET in pd.ExcelFile(XLSX).sheet_names else 0, header=0)
    has_stat = data[~data['Stat Value'].isna()]
    no_rr = has_stat[has_stat['RR Stat Value'].isna()]
    dep = data[data['Type'].isin(['dependency_priors', 'dependency_distal'])]
    missing_idx = dep[dep['index2'].isna() & (dep['Type'] != 'dependency_distal')]  # dependency_distal has no index2
    has_rr = data[~data['RR Stat Value'].isna()]
    no_pm = has_rr[has_rr['RR plus minus'].isna()]

    for name, val in [('Stat no RR', len(no_rr)), ('Missing idx2', len(missing_idx)), ('RR no pm', len(no_pm))]:
        status = 'PASS' if val == 0 else 'FAIL'
        print(f'  {name}: {val} {status}')
        if val != 0: issues += 1

    if not USE_CSV:
        # File size. Post-formula-flatten the xlsx is ~870 KB (formulas replaced
        # with cached numeric values); a file much smaller than that is suspicious.
        fsize = os.path.getsize(XLSX)
        ok = fsize > 500000 if is_default_xlsx else True
        print(f'  File size: {fsize:,} {"PASS" if ok else "FAIL (too small)"}')
        if not ok: issues += 1

    print(f'  Rows: {len(data)}')

    # Count DD, naive_0, and CPT definitions
    pure_dd_types = {'discrete_priors', 'discrete_nhanes_explicit', 'discrete_nhanes_quartile',
                     'discrete_nhanes_quartile_average'}
    naive_types = {'naive_0_nhanes_explicit', 'naive_0_nhanes_explicit_average',
                   'naive_0_nhanes_quartile', 'naive_0_nhanes_quartile_average'}
    cpt_types = {'dependency_priors', 'dependency_distal', 'dependency_nhanes_explicit', 'dependency_nhanes_quartile',
                 'dependency_nhanes_explicit_average', 'dependency_nhanes_quartile_average',
                 'any_of', 'all_of', 'avg', 'if_then_else'}
    n_pure_dd = len(data[data['Type'].isin(pure_dd_types)])
    n_naive = len(data[data['Type'].isin(naive_types)])
    n_cpt = len(data[data['Type'].isin(cpt_types)])

    # Of naive_0: connected (referenced as input) vs isolated
    naive_nodes = set(data[data['Type'].isin(naive_types)]['output'].unique())
    all_inputs = set(data['input'].dropna().unique())
    n_naive_connected = len(naive_nodes & all_inputs)
    n_naive_isolated = len(naive_nodes - all_inputs)

    # Check if naive_0 is off (dependency.py has `if True:` guard)
    naive_0_off = True  # TODO: detect from code
    try:
        from sn_bayes.config_creation.dependency import parse_dependency
        import inspect
        src = inspect.getsource(parse_dependency)
        naive_0_off = 'if True:' in src and 'naive_0 disabled' in src
    except:
        pass

    if naive_0_off:
        n_dd = n_pure_dd + n_naive_connected
        print(f'  naive_0: OFF')
        print(f'  Spreadsheet: {n_pure_dd} pure DD + {n_naive_connected} connected naive_0 = {n_dd} DD, {n_cpt} CPT, {n_naive_isolated} isolated naive_0 (ignored)')
        # Some pure DD nodes only serve as naive_0 parents. With naive_0 off,
        # those parent links are disabled, so these DDs have no children and are
        # correctly excluded by prepare_config. Count them.
        dd_node_set = set(data[data['Type'].isin(pure_dd_types)]['output'].unique())
        dd_only_naive_children = set()
        for node in dd_node_set:
            children = data[data['input'] == node]['output'].unique()
            if len(children) == 0: continue
            all_naive = all(
                any(t in naive_types for t in data[data['output'] == child]['Type'].dropna().unique())
                for child in children
            )
            if all_naive:
                dd_only_naive_children.add(node)

        n_dd_excluded = len(dd_only_naive_children)
        n_dd_expected = n_dd - n_dd_excluded
        print(f'  Pure DD with only naive_0 children (excluded): {n_dd_excluded}')
        print(f'  Expected pickle: {n_dd_expected} DD, {n_cpt} CPT, total {n_dd_expected + n_cpt}')
        print(f'  Breakdown: ({n_pure_dd} pure DD - {n_dd_excluded} excluded + {n_naive_connected} connected naive_0) = {n_dd_expected} DD')

        # Rule: every naive_0 node with at least one non-naive child must be in the net
        non_naive_def_types = pure_dd_types | cpt_types
        naive_with_nonnaive_children = set()
        for node in naive_nodes:
            children = data[data['input'] == node]['output'].unique()
            for child in children:
                child_types = set(data[data['output'] == child]['Type'].dropna().unique())
                if child_types & non_naive_def_types:
                    naive_with_nonnaive_children.add(node)
                    break
        # These must all be in the connected set
        missing_naive = naive_with_nonnaive_children - (naive_nodes & all_inputs)
        if missing_naive:
            print(f'  FAIL: {len(missing_naive)} naive_0 nodes have non-naive children but are not connected: {sorted(missing_naive)}')
            issues += 1
        else:
            print(f'  Naive_0 with non-naive children: {len(naive_with_nonnaive_children)} (all connected) PASS')
    else:
        # naive_0 ON: connected naive become CPTs, isolated should NOT exist
        n_dd = n_pure_dd
        n_cpt_with_naive = n_cpt + n_naive_connected
        print(f'  naive_0: ON')
        print(f'  Spreadsheet: {n_pure_dd} DD, {n_cpt} CPT + {n_naive_connected} naive_0→CPT = {n_cpt_with_naive} CPT')
        print(f'  Expected pickle: {n_dd} DD, {n_cpt_with_naive} CPT, total {n_dd + n_cpt_with_naive}')
        if n_naive_isolated > 0:
            print(f'  FAIL: {n_naive_isolated} isolated naive_0 nodes - they have parents but nothing references them')
            issues += 1

    # Every study has a citation
    studies = data[~data['Stat Value'].isna()]
    no_cite = studies[studies['citation'].isna()]
    status = 'PASS' if len(no_cite) == 0 else 'FAIL'
    print(f'  Studies without citation: {len(no_cite)} {status}')
    if len(no_cite) > 0: issues += 1

    # Connecting study completeness (dependency_distal inputs with RR)
    if 'verification_status' in data.columns:
        placeholders = data[data['verification_status'] == 'placeholder']
        if len(placeholders) > 0:
            print(f'  Placeholder connecting studies: {len(placeholders)} WARN - need literature')
            for _, r in placeholders.iterrows():
                print(f'    {r["output"]} <- {r["input"]} (RR={r.get("RR Stat Value", "?")})')
        dep_distal_nodes = set(data[data['Type'] == 'dependency_distal']['output'])
        for node in dep_distal_nodes:
            inputs_with_rr = data[(data['output'] == node) & (~data['RR Stat Value'].isna())]
            for _, r in inputs_with_rr.iterrows():
                missing = []
                if pd.isna(r.get('study_n')): missing.append('study_n')
                if pd.isna(r.get('study_design')): missing.append('study_design')
                if pd.isna(r.get('study_population')): missing.append('study_population')
                vs = r.get('verification_status', '')
                # Accept any status that starts with a recognized keyword
                # (case-insensitive) - so 'VERIFIED 2026-04-18: ...' or
                # 'UNVERIFIED - Claude-derived from X' both count as audited.
                # Blank or unrecognized → missing.
                vs_str = str(vs).strip().lower() if not pd.isna(vs) else ''
                if not any(vs_str.startswith(kw) for kw in
                           ('verified', 'unverified', 'placeholder', 'fixed',
                            'check', 'population-limited')):
                    missing.append('verification_status')
                if missing:
                    print(f'  Connecting study {node}<-{r["input"]} missing: {", ".join(missing)} WARN')

    # Stat='rr' rows must have both RR Stat Value and RR plus minus populated.
    # No silent placeholder - every declared RR row carries a number + uncertainty.
    rr_rows = data[data['Stat'].fillna('').str.lower() == 'rr']
    rr_no_value = rr_rows[rr_rows['RR Stat Value'].isna()]
    rr_no_pm = rr_rows[rr_rows['RR plus minus'].isna()]
    n_rr_bad = len(rr_no_value) + len(rr_no_pm)
    rr_status = 'PASS' if n_rr_bad == 0 else 'FAIL'
    print(f'  Stat=rr rows with missing RR/PM: {n_rr_bad} {rr_status}')
    if n_rr_bad > 0:
        for _, r in pd.concat([rr_no_value, rr_no_pm]).drop_duplicates().iterrows():
            print(f'    row {r.name+2}: {r["output"]} <- {r["input"]} | RR={r["RR Stat Value"]!r}, PM={r["RR plus minus"]!r}')
        issues += 1

    # K ≤ 5 per dependency node (manual §0.2: hard rule - inversions spike at K=6+)
    # Known waivers (documented exceptions):
    #   coronary_artery_disease K=6 - biologically K cannot be cut without
    #     dropping one of the 6 established CAD risk factors; we accept
    #     the modest inversion risk for this single node.
    # Add more via env var DATA_CHECK_ALLOW_K6=<comma-separated node names>.
    dep_types = {'dependency_priors', 'dependency_distal', 'dependency_nhanes_explicit', 'dependency_nhanes_quartile',
                 'dependency_nhanes_explicit_average', 'dependency_nhanes_quartile_average'}
    k_waivers = {'coronary_artery_disease'}
    if os.environ.get('DATA_CHECK_ALLOW_K6'):
        k_waivers |= {s.strip() for s in os.environ['DATA_CHECK_ALLOW_K6'].split(',') if s.strip()}
    k_violations = []
    k_waived = []
    for _, r in data[data['Type'].isin(dep_types)].iterrows():
        node = r['output']
        # K_DAG = unique parents (CPT cell count is 4^K_unique). Multi-value rows
        # on the same parent (e.g., bmi_naive at obese + overweight) count as ONE.
        k = data[(data['output'] == node) & (~data['input'].isna())]['input'].nunique()
        if k > 7:  # May-1 directive: K up to 7 permitted (was K>6); avoids K-restore mistake
            entry = f'{node} K={k}'
            if node in k_waivers:
                k_waived.append(entry)
            else:
                k_violations.append(entry)
    if k_violations:
        print(f'  K > 7 violations: {len(k_violations)} FAIL - {", ".join(k_violations)}')
        issues += 1
    elif k_waived:
        print(f'  K > 7 violations: 0 (+{len(k_waived)} waived) PASS')
    else:
        print(f'  K > 7 violations: 0 PASS')

    # κ check: report computed Bonferroni multiplier from current N
    try:
        from sn_bayes.kappa import compute_kappa_for_relations
        kappa, n_studies = compute_kappa_for_relations(data)
        print(f'  Bonferroni κ (computed): {kappa:.4f} from N={n_studies} studies (joint-95%)')
    except Exception as e:
        print(f'  Bonferroni κ: SKIP ({e})')

    # Mixed direction in gates (manual §0.9: directly causes FLAT, not a warning)
    # Known waivers (documented exceptions):
    #   frailty_weight (4r+1p)       - weight gain protective vs. weight loss risk
    #                                   in elderly is literature-established;
    #                                   splitting the gate hides the clinical
    #                                   concept "abnormal weight change".
    #   frailty_healthy_diet (1r+1p) - same reasoning for diet direction.
    # Add more via env var DATA_CHECK_ALLOW_MIXED=<comma-separated gate names>.
    mixed_waivers = {'frailty_weight', 'frailty_healthy_diet'}
    if os.environ.get('DATA_CHECK_ALLOW_MIXED'):
        mixed_waivers |= {s.strip() for s in os.environ['DATA_CHECK_ALLOW_MIXED'].split(',') if s.strip()}
    mixed = []
    mixed_waived = []
    for _, r in data[data['Type'].isin(['any_of', 'all_of'])].iterrows():
        gate = r['output']
        rrs = data[(data['output'] == gate) & (~data['input'].isna())]['RR Stat Value'].dropna()
        rrs = pd.to_numeric(rrs, errors='coerce').dropna()  # drop SMD_UNRELIABLE etc.
        if len(rrs) > 1:
            risk = (rrs > 1.0).sum()
            prot = (rrs < 1.0).sum()
            if risk > 0 and prot > 0:
                entry = f'{gate}({risk}r+{prot}p)'
                if gate in mixed_waivers:
                    mixed_waived.append(entry)
                else:
                    mixed.append(entry)
    if mixed:
        print(f'  Mixed direction gates: {len(mixed)} FAIL - {", ".join(mixed[:5])}')
        issues += 1
    elif mixed_waived:
        print(f'  Mixed direction gates: 0 (+{len(mixed_waived)} waived) PASS')
    else:
        print(f'  Mixed direction gates: 0 PASS')

    # Scattered blocks (manual §2 convention: rows for the same node are
    # contiguous, definition row is last in block). Detect definition rows
    # that are far away from their own input rows - that's a sign someone
    # appended at the end instead of inserting at the logical position.
    # Known waivers: anomaly nodes intentionally pull inputs from far away
    # (they reference leaves throughout the sheet by design).
    SCATTER_WAIVERS = {
        'workout_anomaly', 'steps_anomaly', 'sleep_anomaly',
        'walking_speed_anomaly', 'heart_rate_anomaly',
    }
    if os.environ.get('DATA_CHECK_ALLOW_SCATTER'):
        SCATTER_WAIVERS |= {s.strip() for s in os.environ['DATA_CHECK_ALLOW_SCATTER'].split(',') if s.strip()}
    scattered = []
    scattered_waived = []
    all_def_types_for_scatter = dep_types | {'any_of', 'all_of', 'avg', 'if_then_else'}
    SCATTER_THRESHOLD = 20  # rows of gap between earliest input and def row
    for _, defrow in data[data['Type'].isin(all_def_types_for_scatter)].iterrows():
        gate = defrow['output']
        if pd.isna(gate): continue
        def_idx = defrow.name  # pandas index
        inputs = data[(data['output'] == gate) & (~data['input'].isna())]
        if len(inputs) == 0:
            continue
        earliest = inputs.index.min()
        gap = def_idx - earliest
        if gap > SCATTER_THRESHOLD:
            entry = f'{gate} (def row {def_idx}, earliest input row {earliest}, gap {gap})'
            if gate in SCATTER_WAIVERS:
                scattered_waived.append(entry)
            else:
                scattered.append(entry)
    if scattered:
        print(f'  Scattered blocks: {len(scattered)} WARN ({len(scattered_waived)} waived)')
        for s in scattered[:10]:
            print(f'    - {s}')
        if len(scattered) > 10:
            print(f'    ... +{len(scattered)-10} more')
    elif scattered_waived:
        print(f'  Scattered blocks: 0 (+{len(scattered_waived)} waived) PASS')
    else:
        print(f'  Scattered blocks: 0 PASS')

    # Orphan inputs
    all_outputs = set(data['output'].dropna().unique())
    all_inputs = set(data['input'].dropna().unique())
    orphans = [i for i in (all_inputs - all_outputs) if i in all_inputs]
    # Filter to non-anomaly
    real_orphans = []
    for inp in orphans:
        uses = data[(data['input'] == inp) & (~data['Type'].isin(['anomaly']))]
        if len(uses) > 0 and inp not in all_outputs:
            real_orphans.append(inp)
    status = 'PASS' if len(real_orphans) == 0 else f'WARN: {real_orphans[:5]}'
    print(f'  Orphan inputs: {len(real_orphans)} {status}')

    # Duplicate definitions
    all_def_types = dep_types | {'any_of', 'all_of', 'avg', 'if_then_else',
                'discrete_priors', 'discrete_nhanes_explicit', 'discrete_nhanes_quartile',
                'discrete_nhanes_quartile_average', 'naive_0_nhanes_explicit',
                'naive_0_nhanes_explicit_average', 'naive_0_nhanes_quartile',
                'naive_0_nhanes_quartile_average'}
    defs = data[data['Type'].isin(all_def_types)]
    dup_defs = defs[defs.duplicated(subset=['output'], keep=False)]
    dup_names = sorted(dup_defs['output'].unique()) if len(dup_defs) > 0 else []
    status = 'PASS' if not dup_names else f'FAIL: {dup_names[:5]}'
    print(f'  Duplicate definitions: {len(dup_names)} {status}')
    if dup_names: issues += 1

    # Definition rows must have the value/index columns filled per type.
    # Caught by the Apr 17 gallbladder_cancer bug: it had value1/index1 but no
    # value2/index2, leading to a degenerate CPT where P=0 at overweight.
    # Every other dependency_nhanes_explicit row had both; gallbladder was unique.
    #
    # Different Types need different columns:
    #   - `any_of`, `all_of`, `avg`, `if_then_else`: gates (deterministic). Need
    #     value1 + value2 (output state names). index1/index2 not required -
    #     gates don't carry priors.
    #   - `dependency_priors`, `dependency_nhanes_explicit`,
    #     `dependency_nhanes_explicit_average`: binary priors. Need all 4.
    #   - `dependency_nhanes_quartile`, `..._quartile_average`: quartiles auto-
    #     computed; only value1 + index1 (base name + NHANES code) needed.
    #   - `dependency_distal`: needs value1 + value2 (binary state names).
    gate_types_with_states = {'any_of', 'all_of', 'avg', 'if_then_else',
                              'dependency_distal'}
    binary_prior_types = {'dependency_priors', 'dependency_nhanes_explicit',
                          'dependency_nhanes_explicit_average'}
    quartile_types = {'dependency_nhanes_quartile',
                      'dependency_nhanes_quartile_average'}
    missing = []
    def field_empty(v):
        return pd.isna(v) or (isinstance(v, str) and v.strip() == '')
    for _, r in data.iterrows():
        typ = r.get('Type')
        out = r.get('output')
        if pd.isna(typ) or pd.isna(out): continue
        required = None
        if typ in gate_types_with_states:
            required = ['value1', 'value2']
        elif typ in binary_prior_types:
            required = ['value1', 'index1', 'value2', 'index2']
        elif typ in quartile_types:
            required = ['value1', 'index1']
        else:
            continue
        for field in required:
            if field_empty(r.get(field)):
                missing.append(f'{out} ({typ} missing {field})')
                break
    if missing:
        print(f'  Definitions with missing value/index: {len(missing)} FAIL')
        for m in missing[:10]:
            print(f'    - {m}')
        if len(missing) > 10:
            print(f'    ... +{len(missing)-10} more')
        issues += 1
    else:
        print(f'  Definitions with missing value/index: 0 PASS')

    # Gates with no inputs
    empty_gates = []
    for _, r in data[data['Type'].isin(['any_of', 'all_of'])].iterrows():
        gate = r['output']
        inputs = data[(data['output'] == gate) & (~data['input'].isna())]
        if len(inputs) == 0: empty_gates.append(gate)
    status = 'PASS' if not empty_gates else f'FAIL: {empty_gates}'
    print(f'  Empty gates: {len(empty_gates)} {status}')
    if empty_gates: issues += 1

    # Self-references
    self_refs = data[data['output'] == data['input']]
    status = 'PASS' if len(self_refs) == 0 else f'FAIL: {self_refs["output"].tolist()}'
    print(f'  Self-references: {len(self_refs)} {status}')
    if len(self_refs) > 0: issues += 1

    # Gate study rows: every row with Type=disease_name must have RR, sens, or be gate trigger
    keywords = {'is_a', 'subsumes', 'equivalent_to', 'equivalent_distal', 'any_of', 'all_of', 'avg', 'if_then_else', 'distal', 'anomaly'} | pure_dd_types | naive_types | cpt_types
    gate_rows = data[(~data['Type'].isna()) & (~data['Type'].isin(keywords)) & (~data['input'].isna())]
    gate_type_set = {'any_of', 'all_of', 'avg', 'dependency_distal'}
    gate_no_stats = []
    for idx, r in gate_rows.iterrows():
        has_rr_val = pd.notna(r.get('RR Stat Value')) or pd.notna(r.get('Stat Value'))
        has_sens_val = pd.notna(r.get('Sensitivity Stat Value'))
        inp = r['input']
        is_trigger = len(data[(data['output'] == inp) & (data['Type'].isin(gate_type_set))]) > 0
        if not has_rr_val and not has_sens_val and not is_trigger:
            gate_no_stats.append(f'Row {idx+2}: {r["output"]} ← {inp} Type={r["Type"]}')
    status = 'PASS' if not gate_no_stats else 'FAIL'
    print(f'  Gate study rows without RR/sens/trigger: {len(gate_no_stats)} {status}')
    if gate_no_stats:
        for g in gate_no_stats[:5]: print(f'    {g}')
        issues += 1

    # Gate study rows must have disease tag in Type column
    gate_missing_tag = []
    for _, def_row in data[data['Type'].isin(['any_of', 'all_of'])].iterrows():
        gate = def_row['output']
        inputs = data[(data['output'] == gate) & (~data['input'].isna())]
        for _, inp in inputs.iterrows():
            if pd.notna(inp.get('RR Stat Value')) and pd.isna(inp.get('Type')):
                gate_missing_tag.append(f'Row {inp.name+2}: {gate} ← {inp["input"]}')
    status = 'PASS' if not gate_missing_tag else 'FAIL'
    print(f'  Gate RR inputs missing disease tag: {len(gate_missing_tag)} {status}')
    if gate_missing_tag:
        for g in gate_missing_tag[:5]: print(f'    {g}')
        issues += 1

    # DAG check: no cycles
    edges = {}
    for _, r in data.iterrows():
        out = r.get('output')
        inp = r.get('input')
        if pd.notna(out) and pd.notna(inp):
            edges.setdefault(out, set()).add(inp)

    def find_cycle(edges):
        WHITE, GRAY, BLACK = 0, 1, 2
        color = {n: WHITE for n in edges}
        for targets in edges.values():
            for t in targets:
                if t not in color: color[t] = WHITE
        path = []
        def dfs(node):
            color[node] = GRAY
            path.append(node)
            for neighbor in edges.get(node, []):
                if color.get(neighbor) == GRAY:
                    cs = path.index(neighbor)
                    return path[cs:] + [neighbor]
                if color.get(neighbor) == WHITE:
                    result = dfs(neighbor)
                    if result: return result
            path.pop()
            color[node] = BLACK
            return None
        import sys
        sys.setrecursionlimit(10000)
        for node in list(color.keys()):
            if color[node] == WHITE:
                cycle = dfs(node)
                if cycle: return cycle
        return None

    cycle = find_cycle(edges)
    if cycle:
        print(f'  DAG check: CYCLE FOUND: {" → ".join(cycle)}')
        issues += 1
    else:
        print(f'  DAG check: PASS (no cycles)')

    # ── Bayesian-correctness checks (May 1, per user request) ────────────────
    # `edges` here is dict {child: set(parents)}.

    # 1. Multi-path correlation detection. When two parents of the same node
    #    share a common ancestor, the gate combines them as if independent
    #    (the QP's default), but they are correlated through the ancestor.
    #    Bayesian-correct fix per manual §3 is dep_distal between the two
    #    parents at the wrapped-sibling level. Flag pairs without that
    #    wrapper for review. Heuristic - overcounts mildly; INFO not FAIL.
    def ancestors_of(node, _cache={}):
        if node in _cache: return _cache[node]
        _cache[node] = set()  # cycle-guard
        anc = set()
        for p in edges.get(node, set()):
            anc.add(p)
            anc |= ancestors_of(p)
        _cache[node] = anc
        return anc

    type_col = data['Type'].astype(str) if 'Type' in data.columns else pd.Series([''] * len(data))
    distal_typed_outputs = set(data.loc[type_col.str.contains('distal', case=False, na=False), 'output'].dropna())

    multipath_uncorrected = 0
    multipath_examples = []
    for child, parents in edges.items():
        plist = sorted(parents)
        for i, p1 in enumerate(plist):
            for p2 in plist[i+1:]:
                shared = (ancestors_of(p1) & ancestors_of(p2)) - {p1, p2, child}
                if shared and (p1 not in distal_typed_outputs and p2 not in distal_typed_outputs):
                    multipath_uncorrected += 1
                    if len(multipath_examples) < 5:
                        sh = sorted(shared)[:2]
                        multipath_examples.append(f'{child} ← ({p1}, {p2}) share {sh}')
    if multipath_uncorrected:
        print(f'  Multi-path correlation candidates (consider dep_distal): {multipath_uncorrected} INFO')
        for ex in multipath_examples:
            print(f'    e.g. {ex}')
    else:
        print(f'  Multi-path correlation candidates: 0 PASS')

    # 2. dep_distal 3-row pattern integrity. Per manual §3 every dep_distal
    #    wrapper has THREE rows: (a) equivalent_distal definition,
    #    (b) dependency_distal main, (c) connecting study from sibling.
    #    Either piece missing without the others is structural rot.
    eq_distal_outputs = set(data.loc[type_col == 'equivalent_distal', 'output'].dropna())
    dep_distal_outputs = set(data.loc[type_col == 'dependency_distal', 'output'].dropna())
    orphan_eq = eq_distal_outputs - dep_distal_outputs
    orphan_dep = dep_distal_outputs - eq_distal_outputs
    n_orphan = len(orphan_eq) + len(orphan_dep)
    status = 'PASS' if n_orphan == 0 else 'WARN'
    print(f'  dep_distal 3-row integrity: {len(eq_distal_outputs)} eq + {len(dep_distal_outputs)} dep, '
          f'{n_orphan} orphan {status}')
    if orphan_eq:
        print(f'    eq_distal orphans (no dep_distal partner): {sorted(orphan_eq)[:3]}')
    if orphan_dep:
        print(f'    dep_distal orphans (no eq_distal partner): {sorted(orphan_dep)[:3]}')

    # 2b. dep_distal pattern conformance. Each new wrapper must match the
    #     pattern from the user's Apr 16 spec:
    #       def row: Type=dependency_distal, value1/value2 populated,
    #                index1/index2 placeholder priors (recalculated at parse
    #                time from the equivalent_distal parent)
    #       eq_distal row: Type=equivalent_distal, input=<original gate>
    #       original (eq_distal's input) must exist as a separate
    #       definition row (any_of/all_of/dependency)
    #     NOTE: the `_wrapped_yes/_no` suffix is a *human-readability
    #     convention* in our existing data, NOT a hard rule. The parser maps
    #     positionally (config_creation/dependency.py:463-466), so other
    #     value naming works equally. Don't enforce the suffix here.
    pattern_failures = []
    for wrapper in sorted(dep_distal_outputs):
        wrapper_rows = data[data['output'] == wrapper]
        def_rows = wrapper_rows[wrapper_rows['Type'] == 'dependency_distal']
        eq_rows = wrapper_rows[wrapper_rows['Type'] == 'equivalent_distal']
        # check def row
        if len(def_rows) != 1:
            pattern_failures.append(f"{wrapper}: expected 1 def row, found {len(def_rows)}")
            continue
        d = def_rows.iloc[0]
        v1, v2, i1, i2 = d.get('value1'), d.get('value2'), d.get('index1'), d.get('index2')
        if pd.isna(v1) or pd.isna(v2):
            pattern_failures.append(f"{wrapper}: def row missing value1/value2")
        if pd.isna(i1) or pd.isna(i2):
            pattern_failures.append(f"{wrapper}: def row missing index1/index2 priors")
        # check eq_distal row
        if len(eq_rows) != 1:
            pattern_failures.append(f"{wrapper}: expected 1 eq_distal row, found {len(eq_rows)}")
            continue
        e = eq_rows.iloc[0]
        eq_input, eq_input_vals = e.get('input'), e.get('input values')
        if pd.isna(eq_input):
            pattern_failures.append(f"{wrapper}: eq_distal row missing input")
        else:
            # eq_distal input must be a defined node (definition row exists)
            input_def = data[(data['output'] == eq_input) &
                             (data['Type'].isin({'any_of', 'all_of', 'dependency',
                                                 'dependency_priors',
                                                 'dependency_nhanes_explicit',
                                                 'dependency_nhanes_quartile',
                                                 'dependency_nhanes_explicit_average',
                                                 'dependency_nhanes_quartile_average',
                                                 'discrete_nhanes_explicit',
                                                 'discrete_nhanes_quartile'}))]
            if input_def.empty:
                pattern_failures.append(
                    f"{wrapper}: eq_distal input={eq_input!r} has no definition row")
        if pd.isna(eq_input_vals):
            pattern_failures.append(f"{wrapper}: eq_distal row missing input values")
    n_pf = len(pattern_failures)
    pf_status = 'PASS' if n_pf == 0 else 'FAIL'
    print(f'  dep_distal pattern conformance: {n_pf} failures {pf_status}')
    for msg in pattern_failures[:10]:
        print(f'    {msg}')
    if n_pf > 0: issues += n_pf

    # 3. Independence-treated correlated parents. For each gate (any_of/
    #    all_of), check whether its parents are NHANES-correlated. This is
    #    a leading indicator that the QP will treat them as independent
    #    when they aren't. Defer the actual ρ lookup to rho_gap_audit.py
    #    (which has NHANES); just flag any_of/all_of gates with K≥2.
    if 'Type' in data.columns:
        agg_gates = data.loc[type_col.isin({'any_of', 'all_of'}), 'output'].dropna()
        n_agg = len(set(agg_gates))
        # Only the count of aggregator gates is meaningful at the structural-
        # check level; the actual NHANES-ρ-vs-network-ρ comparison is done
        # by scripts/rho_gap_audit.py (which is part of the 5-core panel).
        print(f'  any_of / all_of gate count (review parent-pair ρ via rho_gap_audit): {n_agg} INFO')

    # Connection count: how many input→output links exist
    connections = data[~data['input'].isna() & ~data['output'].isna()]
    n_connections = len(connections)
    unique_links = connections[['output', 'input']].drop_duplicates()
    print(f'  Connections: {n_connections} rows, {len(unique_links)} unique links')

    # Compare with previous build
    prev = load_previous_counts()
    if prev and 'spreadsheet' in prev:
        p = prev['spreadsheet']
        dd_changed = p.get('dd', 0) != n_dd
        cpt_changed = p.get('cpt', 0) != n_cpt
        if dd_changed or cpt_changed:
            print(f'  vs previous build: DD {p.get("dd",0)}→{n_dd}, CPT {p.get("cpt",0)}→{n_cpt} CHANGED')
        else:
            print(f'  vs previous build: DD={n_dd}, CPT={n_cpt} SAME')

    print(f'\n  {"ALL PASSED" if issues == 0 else f"{issues} ISSUES FOUND"}')
    return issues == 0


def check_config(config_path):
    """Run after config creation, before build."""
    print('=' * 60)
    print('CHECK 2: CONFIG (before build)')
    print('=' * 60)
    issues = 0

    with open(config_path) as f:
        config = json.load(f)

    n_nodes = len(config['dependency_data'])
    dd_cfg_types = {'discrete', 'discrete_priors', 'discrete_nhanes_explicit', 'discrete_nhanes_quartile',
                    'discrete_nhanes_quartile_average'}
    naive_cfg_types = {'naive_0_nhanes_explicit', 'naive_0_nhanes_explicit_average',
                       'naive_0_nhanes_quartile', 'naive_0_nhanes_quartile_average'}
    cpt_cfg_types = {'dependency', 'dependency_priors', 'dependency_distal', 'dependency_nhanes_explicit', 'dependency_nhanes_quartile',
                     'dependency_nhanes_explicit_average', 'dependency_nhanes_quartile_average',
                     'any_of', 'all_of', 'avg', 'if_then_else',
                     'is_a', 'subsumes', 'equivalent_to'}
    n_dd = sum(1 for nd in config['dependency_data'].values() if nd.get('TYPE') in dd_cfg_types)
    n_naive = sum(1 for nd in config['dependency_data'].values() if nd.get('TYPE') in naive_cfg_types)
    n_cpt = sum(1 for nd in config['dependency_data'].values() if nd.get('TYPE') in cpt_cfg_types)
    print(f'  Config: {n_dd} DD + {n_naive} naive_0 = {n_dd + n_naive} DD total, {n_cpt} CPT, total {n_nodes}')

    # NaN/None
    nan_c = none_c = 0
    def chk(obj):
        nonlocal nan_c, none_c
        if obj is None: none_c += 1
        elif isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)): nan_c += 1
        elif isinstance(obj, dict):
            for v in obj.values(): chk(v)
        elif isinstance(obj, list):
            for v in obj: chk(v)
    chk(config)

    for name, val in [('NaN/Inf', nan_c), ('None', none_c)]:
        status = 'PASS' if val == 0 else 'FAIL'
        print(f'  {name}: {val} {status}')
        if val != 0: issues += 1

    # RR without pm
    pm_miss = 0
    for name, nd in config['dependency_data'].items():
        inv = nd.get('INVARS', {})
        if not isinstance(inv, dict): continue
        for p, es in inv.items():
            for e in es:
                s = e.get('STATS', {})
                if s and s.get('relative_risk') is not None and s.get('plus_minus') is None:
                    pm_miss += 1
                    print(f'    RR w/o pm: {name} <- {p}')

    status = 'PASS' if pm_miss == 0 else 'FAIL'
    print(f'  RR without pm: {pm_miss} {status}')
    if pm_miss != 0: issues += 1

    # Every dependency INVARS entry must have RR or sensitivity
    dep_cfg_types = {'dependency', 'dependency_priors', 'dependency_distal', 'dependency_nhanes_explicit', 'dependency_nhanes_quartile',
                     'dependency_nhanes_explicit_average', 'dependency_nhanes_quartile_average'}
    no_stats = []
    cfg_rr = 0
    cfg_sens = 0
    for name, nd in config['dependency_data'].items():
        if nd.get('TYPE') not in dep_cfg_types: continue
        inv = nd.get('INVARS', {})
        if not isinstance(inv, dict): continue
        for parent, entries in inv.items():
            for e in entries:
                stats = e.get('STATS', {})
                if 'relative_risk' in stats:
                    cfg_rr += 1
                elif 'sensitivity' in stats:
                    cfg_sens += 1
                else:
                    no_stats.append(f'{name} ← {parent}')
    status = 'PASS' if len(no_stats) == 0 else 'FAIL'
    print(f'  Dependency inputs: {cfg_rr} RR, {cfg_sens} sens/spec, {len(no_stats)} missing {status}')
    if no_stats:
        for s in no_stats[:5]: print(f'    {s}')
        issues += 1

    # ── Bayesian-correctness: dependency nodes must have at least 1 parent (May 1) ──
    # A dependency-typed node with empty INVARS has nothing to condition on - the
    # CPT shape becomes (n_outvars,) instead of (n_parents..., n_outvars), which
    # crashes pomegranate's ConditionalCategorical. Either the upstream patches
    # blanked all parents, or the node was a target without ever having parents.
    parentless_deps = []
    for name, nd in config['dependency_data'].items():
        if nd.get('TYPE') not in dep_cfg_types: continue
        inv = nd.get('INVARS', {})
        if not inv:
            parentless_deps.append(name)
    status = 'PASS' if len(parentless_deps) == 0 else 'FAIL'
    print(f'  Dependency nodes with 0 parents (would crash build): {len(parentless_deps)} {status}')
    if parentless_deps:
        for n in parentless_deps[:10]:
            print(f'    {n}')
        issues += 1

    # Connection count: every INVARS parent must exist as a node
    n_connections = 0
    missing_parents = []
    for name, nd in config['dependency_data'].items():
        inv = nd.get('INVARS', {})
        if not isinstance(inv, dict): continue
        for parent in inv:
            n_connections += 1
            if parent not in config['dependency_data']:
                missing_parents.append(f'{name} ← {parent}')
    print(f'  Connections: {n_connections}')
    if missing_parents:
        print(f'  Missing parents ({len(missing_parents)}):')
        for m in missing_parents[:10]:
            print(f'    {m}')
        if len(missing_parents) > 10: print(f'    ... +{len(missing_parents)-10} more')

    # Nodes in spreadsheet but not in config
    ss_data = pd.read_excel(XLSX, sheet_name=SHEET if SHEET in pd.ExcelFile(XLSX).sheet_names else 0, header=0)
    pure_dd_types = {'discrete_priors', 'discrete_nhanes_explicit', 'discrete_nhanes_quartile', 'discrete_nhanes_quartile_average'}
    naive_types_ss = {'naive_0_nhanes_explicit', 'naive_0_nhanes_explicit_average', 'naive_0_nhanes_quartile', 'naive_0_nhanes_quartile_average'}
    cpt_types_ss = {'dependency_priors', 'dependency_distal', 'dependency_nhanes_explicit', 'dependency_nhanes_quartile',
                 'dependency_nhanes_explicit_average', 'dependency_nhanes_quartile_average',
                 'any_of', 'all_of', 'avg', 'if_then_else'}
    ss_dd = set(ss_data[ss_data['Type'].isin(pure_dd_types)]['output'].unique())
    ss_naive = set(ss_data[ss_data['Type'].isin(naive_types_ss)]['output'].unique())
    ss_cpt = set(ss_data[ss_data['Type'].isin(cpt_types_ss)]['output'].unique())
    ss_all_inputs = set(ss_data['input'].dropna().unique())
    ss_connected = (ss_dd | (ss_naive & ss_all_inputs) | ss_cpt)

    dropped = ss_connected - set(config['dependency_data'].keys())
    if dropped:
        print(f'  Spreadsheet nodes dropped from config ({len(dropped)}):')
        for d in sorted(dropped):
            refs = ss_data[ss_data['input'] == d]['output'].unique()
            print(f'    {d} (referenced by {len(refs)} nodes)')

    # ── Bayesian-correctness: OUTVARS shape consistency (May 1) ──────────
    # OUTVARS shows up as dict {value_name: probability} (linearized form)
    # OR list [name, prob, name, prob, ...] (legacy form) OR list [name, name]
    # (just outvar names, no probs). Code in distal.py / dependency.py /
    # utils.py has different expectations across these shapes - see
    # docs/FIXES_LEDGER.md for the bug history. Verify each OUTVARS is
    # well-formed and probabilities sum to 1 when present.
    outvars_bad_shape = 0
    outvars_bad_sum = 0
    examples = []
    for name, nd in config['dependency_data'].items():
        ov = nd.get('OUTVARS')
        if ov is None:
            continue
        if isinstance(ov, dict):
            if not all(isinstance(v, (int, float)) for v in ov.values()):
                outvars_bad_shape += 1
                if len(examples) < 3:
                    examples.append(f'{name}: dict-OUTVARS but non-numeric values')
                continue
            total = sum(ov.values())
            if abs(total - 1.0) > 0.01:
                outvars_bad_sum += 1
                if len(examples) < 3:
                    examples.append(f'{name}: dict-OUTVARS sums to {total:.4f} (not 1)')
        elif isinstance(ov, list):
            # Either [name, name, ...] (str-only) or [name, p, name, p, ...]
            # mixed. Both are legitimate legacy forms.
            if len(ov) == 0:
                outvars_bad_shape += 1
                if len(examples) < 3:
                    examples.append(f'{name}: empty list-OUTVARS')
            # Else accepted; can't verify sum without parsing form
        else:
            outvars_bad_shape += 1
            if len(examples) < 3:
                examples.append(f'{name}: OUTVARS type={type(ov).__name__} (expected dict or list)')

    status = 'PASS' if outvars_bad_shape + outvars_bad_sum == 0 else 'FAIL'
    print(f'  OUTVARS shape consistency: {outvars_bad_shape} bad shape, {outvars_bad_sum} bad sum {status}')
    for ex in examples:
        print(f'    {ex}')
    if outvars_bad_shape + outvars_bad_sum:
        issues += 1

    # ── Bayesian-correctness: dependency_distal wrapper integrity (May 1) ──
    # A dep_distal node should have an INVARS containing the wrapped node
    # (via equivalent_distal type) plus at least one connecting study from
    # a correlated sibling. Empty INVARS = wrapper has no connection.
    dep_distal_empty = 0
    for name, nd in config['dependency_data'].items():
        if nd.get('TYPE') != 'dependency_distal':
            continue
        inv = nd.get('INVARS', {})
        if not inv or len(inv) == 0:
            dep_distal_empty += 1
    status = 'PASS' if dep_distal_empty == 0 else 'WARN'
    print(f'  dependency_distal wrappers with empty INVARS: {dep_distal_empty} {status}')

    print(f'\n  {"ALL PASSED" if issues == 0 else f"{issues} ISSUES FOUND"}')
    return issues == 0


def check_pickle(pickle_path):
    """Run after build, before query."""
    print('=' * 60)
    print('CHECK 3: PICKLE (before query)')
    print('=' * 60)
    issues = 0

    proto = smart_load_pickle(pickle_path)

    dds = len(proto.discreteDistributions)
    cpts = len(proto.conditionalProbabilityTables)
    print(f'  DDs: {dds}, CPTs: {cpts}, Pickle total: {dds + cpts}')

    # Empty CPTs
    empty = [c.name for c in proto.conditionalProbabilityTables if len(c.conditionalProbabilityRows) == 0]
    status = 'PASS' if len(empty) == 0 else 'FAIL'
    print(f'  Empty CPTs: {len(empty)} {status}')
    if empty:
        print(f'    {empty}')
        issues += 1

    # NaN/negative/>1 in CPTs. Negative threshold tightened from <-0.01 to <0:
    # the Apr 17 gallbladder_cancer bug produced P=-0.000103 (within old
    # tolerance) that the network engine clamped to 0, causing a silent
    # direction flip. Any negative probability is invalid - flag it.
    nan_nodes = set()
    neg_nodes = set()
    over_nodes = set()
    neg_examples = []  # (node, parent_vals, probability)
    for c in proto.conditionalProbabilityTables:
        for r in c.conditionalProbabilityRows:
            if np.isnan(r.probability): nan_nodes.add(c.name)
            if r.probability < 0:
                neg_nodes.add(c.name)
                if len(neg_examples) < 5:
                    vals = [v.name for v in r.randomVariableValues]
                    neg_examples.append((c.name, vals, r.probability))
            if r.probability > 1.0 + 1e-6: over_nodes.add(c.name)

    for name, nodes in [('NaN CPTs', nan_nodes), ('Negative CPTs', neg_nodes), ('>1 CPTs', over_nodes)]:
        status = 'PASS' if len(nodes) == 0 else 'FAIL'
        print(f'  {name}: {len(nodes)} {status}')
        if nodes:
            print(f'    {sorted(nodes)}')
            issues += 1
    if neg_examples:
        print('  Negative CPT cell examples (first 5):')
        for node, vals, p in neg_examples:
            print(f'    {node}: {vals} P={p:.6f}')

    # DD sums
    dd_bad = 0
    for dd in proto.discreteDistributions:
        probs = [v.probability for v in dd.variables]
        if abs(sum(probs) - 1.0) > 0.01:
            dd_bad += 1
        for v in dd.variables:
            if np.isnan(v.probability) or v.probability < 0 or v.probability > 1:
                dd_bad += 1

    status = 'PASS' if dd_bad == 0 else 'FAIL'
    print(f'  DD issues: {dd_bad} {status}')
    if dd_bad != 0: issues += 1

    print(f'  Anomalies: {len(proto.anomalies)}')

    # Connection count: parents per CPT
    n_connections = 0
    for cpt in proto.conditionalProbabilityTables:
        n_connections += len(cpt.randomVariables)
    print(f'  CPT parent connections: {n_connections}')

    # Check all CPT parents exist as DD or CPT
    all_names = set(dd.name for dd in proto.discreteDistributions) | set(cpt.name for cpt in proto.conditionalProbabilityTables)
    missing = []
    for cpt in proto.conditionalProbabilityTables:
        for rv in cpt.randomVariables:
            if rv.name not in all_names:
                missing.append(f'{cpt.name} ← {rv.name}')
    if missing:
        print(f'  CPT parents not in pickle ({len(missing)}):')
        for m in missing[:10]:
            print(f'    {m}')
        issues += 1

    # ── Bayesian-correctness: CPT row sums must equal 1 (May 1) ──────────
    # For each CPT, group rows by parent assignment; the probabilities of
    # the child's values conditional on the parent assignment must sum to 1.
    # Probability axiom - silent violations indicate solver bugs.
    n_bad_rowsum = 0
    bad_rowsum_examples = []
    for cpt in proto.conditionalProbabilityTables:
        groups = {}
        for r in cpt.conditionalProbabilityRows:
            # The parent assignment is all randomVariableValues except the last
            # (which is the child's outcome). Group by tuple of parent values.
            parent_assignment = tuple(v.name for v in r.randomVariableValues[:-1])
            groups.setdefault(parent_assignment, 0.0)
            groups[parent_assignment] += r.probability
        for assignment, total in groups.items():
            if abs(total - 1.0) > 0.01:
                n_bad_rowsum += 1
                if len(bad_rowsum_examples) < 5:
                    bad_rowsum_examples.append(f'{cpt.name} | parents={assignment} sum={total:.4f}')
                break  # one bad row per CPT is enough to flag

    status = 'PASS' if n_bad_rowsum == 0 else 'FAIL'
    print(f'  CPT row sums = 1 (Bayesian axiom): {n_bad_rowsum} violations {status}')
    for ex in bad_rowsum_examples:
        print(f'    {ex}')
    if n_bad_rowsum:
        issues += 1

    # ── Bayesian-correctness: each CPT row count ≤ ∏(parent cardinalities) (May 1) ──
    # Each CPT must enumerate every (parent_value × child_value) combination
    # exactly once. Counting that the actual row count matches the expected
    # count catches silent missing/duplicate rows.
    n_count_mismatch = 0
    count_mismatch_examples = []
    for cpt in proto.conditionalProbabilityTables:
        # Need parent cardinalities; pull from DDs and other CPTs.
        parent_card = []
        for rv in cpt.randomVariables:
            if rv.name == cpt.name:
                continue
            # Look up in DDs first
            dd_card = None
            for dd in proto.discreteDistributions:
                if dd.name == rv.name:
                    dd_card = len(dd.variables)
                    break
            if dd_card is not None:
                parent_card.append(dd_card)
                continue
            # Else look up CPT outvars
            cpt_card = None
            for c2 in proto.conditionalProbabilityTables:
                if c2.name == rv.name:
                    cpt_card = len(c2.outvars)
                    break
            if cpt_card is not None and cpt_card > 0:
                parent_card.append(cpt_card)
        # Expected rows = ∏ parent_card × n_outvars
        if parent_card and len(cpt.outvars) > 0:
            expected = 1
            for c in parent_card: expected *= c
            expected *= len(cpt.outvars)
            actual = len(cpt.conditionalProbabilityRows)
            if expected != actual and abs(expected - actual) > 0:
                n_count_mismatch += 1
                if len(count_mismatch_examples) < 5:
                    count_mismatch_examples.append(
                        f'{cpt.name}: rows={actual}, expected={expected} '
                        f'(parents={parent_card}, outvars={len(cpt.outvars)})')

    status = 'PASS' if n_count_mismatch == 0 else 'WARN'  # WARN not FAIL - some types have variable structure
    print(f'  CPT row count = ∏(parents) × outvars: {n_count_mismatch} mismatches {status}')
    for ex in count_mismatch_examples:
        print(f'    {ex}')

    print(f'\n  {"ALL PASSED" if issues == 0 else f"{issues} ISSUES FOUND"}')
    return issues == 0


def check_post_run(pickle_path, config_path):
    """Run after full build+query. Checks predictions are sane."""
    from sn_bayes.utils import bayesInitialize, query

    print('=' * 60)
    print('CHECK 4: POST-RUN (after query)')
    print('=' * 60)
    issues = 0

    proto = smart_load_pickle(pickle_path)
    with open(config_path) as f:
        config = json.load(f)

    net = bayesInitialize(proto)

    # Node-count consistency: catches subnet config used with full pickle (or vice versa).
    # Without this check, downstream calibration/AUC/fidelity counts silently undercount.
    proto_node_count = len(set(rv.name for tbl in proto.conditionalProbabilityTables
                                for rv in tbl.randomVariables))
    config_node_count = len(config.get('dependency_data', {}))
    ratio = config_node_count / proto_node_count if proto_node_count else 0
    status = 'PASS' if ratio >= 0.9 else 'FAIL'
    print(f'  Pickle nodes: {proto_node_count}, Config dependency_data: {config_node_count} {status}')
    if ratio < 0.9:
        print(f'    Config covers only {ratio:.0%} of pickle nodes; metrics will undercount.')
        issues += 1

    # Baseline query - check for NaN
    pp = net.predict_proba({})
    nan_dists = 0
    if isinstance(pp, dict):
        for name, dist in pp.items():
            if isinstance(dist, dict):
                for v in dist.values():
                    if np.isnan(float(v)):
                        nan_dists += 1
                        break

    status = 'PASS' if nan_dists == 0 else 'FAIL'
    print(f'  NaN in predict_proba: {nan_dists} {status}')
    if nan_dists != 0: issues += 1

    # Key disease baselines
    diseases = ['diabetes', 'cancer', 'cardiovascular_disease', 'frailty',
                'all_cause_mortality', 'stroke', 'cognitive_impairment']
    r0 = query(net, proto, {}, diseases)
    print(f'  Baseline probabilities:')
    for d in diseases:
        if d in r0:
            pk = list(r0[d].keys())[0]
            p = float(r0[d][pk])
            flag = ' WARN: NaN' if np.isnan(p) else ' WARN: 0' if p == 0 else ' WARN: 1' if p >= 1 else ''
            print(f'    {pk:50s} = {p:.4f}{flag}')
            if np.isnan(p) or p == 0 or p >= 1: issues += 1

    # Validation windows summary
    import glob
    val_files = sorted(glob.glob('bayesnet_initialize_output/*_validation.csv'))
    if val_files:
        dfs = []
        for f in val_files:
            try: dfs.append(pd.read_csv(f))
            except: pass
        if dfs:
            val = pd.concat(dfs, ignore_index=True)
            w = val['window']
            print(f'  Windows: {len(val)} rows, mean={w.mean():.4f}, median={w.median():.4f}, max={w.max():.4f}')

    print(f'\n  {"ALL PASSED" if issues == 0 else f"{issues} ISSUES FOUND"}')
    return issues == 0


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(f'Usage: {sys.argv[0]} spreadsheet|config|pickle|post_run [args]')
        print(f'  --xlsx <path>   Override spreadsheet path')
        print(f'  --sheet <name>  Override sheet name')
        print(f'  --build-dir <path>  Override build directory')
        sys.exit(1)

    stage = sys.argv[1]
    build_dir = None
    # Parse optional arguments
    for i, arg in enumerate(sys.argv):
        if arg == '--build-dir' and i + 1 < len(sys.argv):
            build_dir = sys.argv[i + 1]
        if arg == '--xlsx' and i + 1 < len(sys.argv):
            XLSX = sys.argv[i + 1]
            USE_CSV = False  # explicit xlsx path turns off CSV
        if arg == '--use-xlsx':
            USE_CSV = False
        if arg == '--sheet' and i + 1 < len(sys.argv):
            SHEET = sys.argv[i + 1]
        if arg == '--csv' and i + 1 < len(sys.argv) and not sys.argv[i + 1].startswith('--'):
            CSV_PATH = sys.argv[i + 1]
            USE_CSV = True
        elif arg == '--csv' or arg == '--use-csv':
            USE_CSV = True

    if stage == 'spreadsheet':
        ok = check_spreadsheet()
    elif stage == 'config':
        ok = check_config(sys.argv[2] if len(sys.argv) > 2 and not sys.argv[2].startswith('--') else 'bayesnet_config_linear.json')
    elif stage == 'pickle':
        ok = check_pickle(sys.argv[2] if len(sys.argv) > 2 and not sys.argv[2].startswith('--') else 'bayesianNetworkProto.pickle')
    elif stage == 'post_run':
        p = sys.argv[2] if len(sys.argv) > 2 and not sys.argv[2].startswith('--') else 'bayesianNetworkProto.pickle'
        c = sys.argv[3] if len(sys.argv) > 3 and not sys.argv[3].startswith('--') else 'bayesnet_config_linear.json'
        ok = check_post_run(p, c)
    elif stage == 'all':
        # Run all 4 checks in sequence, save counts to build dir
        ok = check_spreadsheet()
        if ok:
            cfg = sys.argv[2] if len(sys.argv) > 2 and not sys.argv[2].startswith('--') else 'bayesnet_config_linear.json'
            ok = check_config(cfg)
        if ok:
            pkl = sys.argv[3] if len(sys.argv) > 3 and not sys.argv[3].startswith('--') else 'bayesianNetworkProto.pickle'
            ok = check_pickle(pkl)
        if ok:
            pkl = sys.argv[3] if len(sys.argv) > 3 and not sys.argv[3].startswith('--') else 'bayesianNetworkProto.pickle'
            cfg = sys.argv[2] if len(sys.argv) > 2 and not sys.argv[2].startswith('--') else 'bayesnet_config_linear.json'
            ok = check_post_run(pkl, cfg)
    else:
        print(f'Unknown stage: {stage}')
        sys.exit(1)

    sys.exit(0 if ok else 1)
